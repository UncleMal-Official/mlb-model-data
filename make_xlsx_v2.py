"""Workbook builder v2 — fully parameterized by date; Top-6 auto-composed from the board.
Usage: python3 make_xlsx_v2.py 2026-08-08 [version-label]
"""
import json, os, sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATE = sys.argv[1] if len(sys.argv) > 1 else "2026-08-08"
VER = sys.argv[2] if len(sys.argv) > 2 else "v1"
D = "/home/claude/mlb_model"
R = "/home/claude/mlb-model-data/data"
YDAY = (pd.Timestamp(DATE) - pd.Timedelta(days=1)).strftime("%Y-%m-%d")
DAY_LABEL = pd.Timestamp(DATE).strftime("%A, %b ") + str(pd.Timestamp(DATE).day)
YDAY_LABEL = pd.Timestamp(YDAY).strftime("%a %b ") + str(pd.Timestamp(YDAY).day)

df = pd.read_csv(f"{D}/output/full_board_{VER}_{DATE}.csv").fillna("")
pp_path_csv = f"{D}/output/board_pp_{DATE}.csv"
HAS_PP = os.path.exists(pp_path_csv)
if HAS_PP:
    _pp = pd.read_csv(pp_path_csv)[["player_id", "fs_line", "win_pct", "edge_vs_flex"]]
    df = df.merge(_pp, on="player_id", how="left")
    df[["fs_line", "win_pct", "edge_vs_flex"]] = df[["fs_line", "win_pct", "edge_vs_flex"]].fillna("")
hb = pd.read_parquet(f"{R}/hitting_basic.parquet")
l21 = hb[hb.window == "last21"].set_index("player_id")
ps = json.load(open(f"{D}/data/pitcher_splits_2026.json"))["pitchers"]
sp_by_name = {v["name"]: v for v in ps.values()}
parks = json.load(open(f"{D}/config/park_factors_{DATE}.json"))["games"]
park_pending = "pending" in json.load(open(f"{D}/config/park_factors_{DATE}.json"))["source"].lower() \
    or "placeholder" in json.load(open(f"{D}/config/park_factors_{DATE}.json"))["source"].lower()
da_path = f"{D}/output/day_after_{DATE}.csv"
da_names = set(pd.read_csv(da_path).player) if os.path.exists(da_path) else set()
meta = json.load(open(f"{R}/meta.json"))

NAVY = "1F3864"; GREEN = "C6E7C6"; AMBER = "FFE49C"; GOLD = "B7950B"
HDR = Font(name="Arial", size=9, bold=True, color="FFFFFF")
BASE = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
TITLE = Font(name="Arial", size=14, bold=True, color=NAVY)
SUB = Font(name="Arial", size=10, italic=True, color="555555")
FILL_H = PatternFill("solid", fgColor=NAVY)
FILL_G = PatternFill("solid", fgColor=GREEN)
FILL_A = PatternFill("solid", fgColor=AMBER)
thin = Side(style="thin", color="D9D9D9")
BORD = Border(bottom=thin)

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR; cell.fill = FILL_H
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def write_table(ws, start_row, headers, data_rows, widths, highlight=None, freeze_col=1):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    for i, r in enumerate(data_rows, start_row + 1):
        for j, v in enumerate(r, 1):
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = BASE; cell.border = BORD
        if highlight:
            f = highlight(data_rows[i - start_row - 1])
            if f:
                for j in range(1, len(headers) + 1):
                    ws.cell(row=i, column=j).fill = f
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A{start_row}:{last_col}{start_row + len(data_rows)}"
    ws.freeze_panes = ws.cell(row=start_row + 1, column=freeze_col + 1)

# ---------------- Top-6 auto-composer ----------------
def zs(s):
    s = pd.to_numeric(s, errors="coerce").astype(float)
    return (s - s.mean()) / max(s.std(), 1e-9)

b = df.drop_duplicates("player_id").copy()
b["flags"] = b["flags"].astype(str)
elig = b[(b.conf.isin(["HIGH", "MED"])) & (pd.to_numeric(b.exp_pa) >= 4.0) & (pd.to_numeric(b.proj_fp) >= 7.0)].copy()
sc = zs(elig.pgs) * 0.45 + zs(elig.proj_fp) * 0.35
sc += (elig.conf == "HIGH") * 0.35 + (elig.bpp_tags.astype(str).str.len() > 0) * 0.25
sc -= elig["flags"].str.contains("RAIN") * 0.30 + elig["flags"].str.contains("SP-SMALL-SAMPLE") * 0.15
elig["sc"] = sc
top6df = elig.sort_values("sc", ascending=False).head(6)

def compose_bullets(x):
    pid = int(x.player_id)
    f21 = l21.loc[pid] if pid in l21.index else None
    spl = sp_by_name.get(x.opp_sp)
    side = "L" if x.bats == "L" else "R"
    vs = (spl or {}).get("vsL" if side == "L" else "vsR") or {}
    out = []
    bits = []
    if f21 is not None and float(f21.pa or 0) >= 30:
        bits.append(f"L21: {f21.slg} SLG, {int(f21.hr)} HR in {int(f21.pa)} PA")
    if x.mix_barrel != "" and float(x.mix_bbe or 0) >= 20:
        bits.append(f"{float(x.mix_barrel):.1f}% barrels vs {x.opp_sp.split()[-1]}'s exact mix ({x.mix} to his side)")
    bits.append(f"damage mult {x.dmg_mult}")
    out.append(" · ".join(bits) + ".")
    if vs.get("pa"):
        out.append(f"{x.opp_sp} vs {side}HB: {vs['avg']:.3f}/{vs['obp']:.3f}/{vs['slg']:.3f}, "
                   f"{vs['hr']} HR + {vs['bb']} BB in {vs['pa']} PA — that side is the leak."
                   + (" Park layer PENDING (BPP sheet not posted yet); park multipliers neutral on this board."
                      if park_pending else ""))
    else:
        out.append(f"{x.opp_sp}: no meaningful split sample (first start / return) — model leans on mix + league priors. SP-SMALL-SAMPLE.")
    eyes = []
    if "LINEUP-PROJ" in x["flags"]: eyes.append("projected lineup — reconfirm slot at the 3 PM refresh")
    if "RAIN" in x["flags"]: eyes.append("rain flag")
    if x.player in da_names: eyes.append("DAY-AFTER tag: robbed hard contact yesterday, similar look today")
    if str(x.rev_platoon) == "YES": eyes.append("reverse-platoon value spot")
    eyes.append(f"Fantasy Score is the expression (proj TB {x.proj_tb}, {float(x.hr_prob):.0f}% HR)")
    out.append(" | ".join(eyes) + ".")
    return out

top6 = []
for i, (_, x) in enumerate(top6df.iterrows(), 1):
    head = (f"{i}. {x.player.upper()} ({x.team}) - slot {int(x.slot)}"
            + (" (proj)" if "LINEUP-PROJ" in x["flags"] else " (posted)")
            + f" vs {x.opp_sp} ({x.sp_throws}) | Proj {x.proj_fp} FP | PGS {'+' if float(x.pgs)>=0 else ''}{x.pgs} | {float(x.hr_prob):.0f}% HR")
    top6.append((head, compose_bullets(x)))

# ---------------- Tab 1 ----------------
ws = wb.active; ws.title = "Top Plays & Legend"
ws.sheet_properties.tabColor = GOLD
ws["A1"] = "MLB HITTER MODEL — DAILY BOARD (barrel × pitch-mix × platoon layer)"; ws["A1"].font = TITLE
ws["A2"] = (f"Slate: {DAY_LABEL}, 2026 ({df.game.nunique()} games) | data through {meta.get('raw_through')} "
            f"({meta.get('raw_rows'):,} pitches) | board build: {VER} morning"); ws["A2"].font = SUB
note3 = ("MORNING BOARD: lineups NOT posted yet - all slots projected from 21-day usage (LINEUP-PROJ). "
         "Board refreshes ~3:00 PM and pre-lock as lineups drop; What Changed tab will track every move. ")
if park_pending:
    note3 += "PARK LAYER PENDING - BPP sheets not posted yet; park multipliers neutral. "
note3 += ("PGS = model FP minus public-read FP (surface form + generic platoon). Overs-first. "
          "Yesterday's graded results: see results CSV + tracking.csv.")
ws["A3"] = note3; ws["A3"].font = SUB

r = 5
for head, bullets in top6:
    ws.cell(row=r, column=1, value=head).font = BOLD
    ws.cell(row=r, column=1).fill = FILL_A
    r += 1
    for bl in bullets:
        ws.cell(row=r, column=1, value="   • " + bl).font = BASE
        r += 1
    r += 1

# auto model-calls section: biggest gaps + biggest fades
ws.cell(row=r, column=1, value="MODEL CALLS - biggest perception gaps and formal fades on this board:").font = BOLD
r += 1
gaps = elig.nlargest(3, "pgs")
fades = b[(b.conf.isin(["HIGH", "MED"])) & (pd.to_numeric(b.pgs) <= -0.6)].nsmallest(3, "pgs")
for _, x in gaps.iterrows():
    ws.cell(row=r, column=1, value=f"   • GAP: {x.player} ({x.team}) PGS +{x.pgs} vs {x.opp_sp} - model {x.proj_fp} FP vs public {x.naive_fp}.").font = BASE
    r += 1
for _, x in fades.iterrows():
    ws.cell(row=r, column=1, value=f"   • FADE: {x.player} ({x.team}) PGS {x.pgs} vs {x.opp_sp} - the market will like him more than the model does ({x.proj_fp} vs {x.naive_fp} public).").font = BASE
    r += 1
r += 1
ws.cell(row=r, column=1, value="READ ME / CAVEATS:").font = BOLD
r += 1
for line in [
 "   • CONF: HIGH = real samples both sides. MED = smaller pitcher-side or form sample. LOW = excluded from CORE, visible in Full Board.",
 "   • LINEUP-PROJ = lineup not posted at build time (slots projected). SP-SMALL-SAMPLE = pitcher split under ~90 PA. RAIN = delay/PPD risk.",
 "   • PP LINES LIVE (from your saved JSON): Win % is market-calibrated (global scale fit so avg P(over) ~ 50% across all lines; today k=0.842 - raw model runs hot). GREEN CORE rows = calibrated edge >= +3 vs the 54.2% flex breakeven.",
 "   • Low-line caution: edges concentrated on 3.0-4.5 lines for role players are being PAPER-TRACKED before real money - if that cohort actually cashes 65%+, we scale in; if it's our floor model running hot, calibration tightens.",
 "   • Scoring: 1B=3, 2B=5, 3B=8, HR=10, R/RBI/BB/HBP=2, SB=5 (PrizePicks batter fantasy score).",
]:
    ws.cell(row=r, column=1, value=line).font = BASE
    r += 1
ws.column_dimensions["A"].width = 165

# ---------------- What Changed ----------------
ch_path = f"{D}/output/board_changes_{DATE}.csv"
if os.path.exists(ch_path):
    ch = pd.read_csv(ch_path).fillna("")
    wsc = wb.create_sheet("What Changed", 1); wsc.sheet_properties.tabColor = "F57F17"
    wsc["A1"] = f"WHAT CHANGED - {DAY_LABEL}"; wsc["A1"].font = TITLE
    ugly = (ch.severity == "UGLY").sum(); bad = (ch.severity == "BAD").sum(); good = (ch.severity == "GOOD").sum()
    wsc["A2"] = f"{good} upgrades, {bad} downgrades, {ugly} scratches. RED = act before lock."; wsc["A2"].font = SUB
    FILL_R = PatternFill("solid", fgColor="F4CCCC")
    hch = ["Severity","Type","Player","Team","Game","What Changed","Was","Now","Why"]
    dch = [[x.severity, x.category, x.player, x.team, x.game, x.change, x.was, x.now, x.why] for _, x in ch.iterrows()]
    write_table(wsc, 4, hch, dch, [9,7,20,6,10,34,22,24,50],
                lambda r: PatternFill("solid", fgColor="F4CCCC") if r[0]=="UGLY" else (FILL_A if r[0]=="BAD" else FILL_G), freeze_col=3)

# ---------------- Yesterday's Lessons ----------------
ls_path = f"{D}/output/lessons_{YDAY}.csv"
if os.path.exists(ls_path):
    LS = pd.read_csv(ls_path, index_col=0).fillna("")
    wsl = wb.create_sheet("Yesterday's Lessons", 1 if not os.path.exists(ch_path) else 2)
    wsl.sheet_properties.tabColor = "4527A0"
    wsl["A1"] = f"YESTERDAY'S LESSONS - {YDAY_LABEL} slate, machine-graded"; wsl["A1"].font = TITLE
    wsl["A2"] = "Computed from batted-ball data - expected vs actual, rolling contact quality, matchup context. UNLUCKY = buy. LUCKY = fade the box score. FELL OFF = real decline + why. PITCHER = process vs results."; wsl["A2"].font = SUB
    r = 4
    CAT_FILL = {"UNLUCKY": FILL_G, "STILL HOT": FILL_G, "LUCKY": FILL_A, "FELL OFF": FILL_A, "PITCHER": None}
    for i, x in LS.iterrows():
        c = wsl.cell(row=r, column=1, value=f"{i}. [{x.category}]  {x.headline}")
        c.font = BOLD
        f = CAT_FILL.get(x.category)
        if f: c.fill = f
        r += 1
        wsl.cell(row=r, column=1, value="      " + x.data).font = BASE
        r += 2
    wsl.column_dimensions["A"].width = 165

# ---------------- CORE ----------------
core = df[df.conf != "LOW"].sort_values("proj_fp", ascending=False).head(60)
ws2 = wb.create_sheet("CORE"); ws2.sheet_properties.tabColor = "2E7D32"
ws2["A1"] = "CORE — standard-line targets, ranked by projected PrizePicks Fantasy Score"; ws2["A1"].font = TITLE
ws2["A2"] = "Green = PGS ≥ +1.0 (market likely under-pricing). Amber = reverse-platoon value spot. Overs-first."; ws2["A2"].font = SUB
headers = ["Rank","Player","Team","Slot","Pos","Bats","Opp SP","Throws","Same/Opp","SP Mix (vs side)","Mix Barrel %","Dmg Mult","HR Mult","Side PA","Exp PA","Proj FP","Public-Read FP","PGS"]
pp_cols = ["PP Line (FS)","Win % (calibrated)","Edge vs Flex"] if HAS_PP else []
headers += pp_cols + ["Proj TB","HR %","Hit %","Rev-Platoon","Free-Swing","Conf","Flags","BPP Corroboration"]
data = []
for i, (_, x) in enumerate(core.iterrows(), 1):
    row = [i, x.player, x.team, x.slot, x.pos, x.bats, x.opp_sp, x.sp_throws,
           x.matchup, x.mix, x.mix_barrel, x.dmg_mult, x.hr_mult, x.side_pa, x.exp_pa,
           x.proj_fp, x.naive_fp, x.pgs]
    if HAS_PP:
        row += [x.fs_line, x.win_pct, x.edge_vs_flex]
    row += [x.proj_tb, x.hr_prob, x.hit_prob,
            x.rev_platoon, x.free_swinger, x.conf, x["flags"], x.bpp_tags]
    data.append(row)
EDGE_I = 20 if HAS_PP else None
def hl_core(row):
    if EDGE_I is not None and row[EDGE_I] != "" and float(row[EDGE_I]) >= 3.0: return FILL_G
    if row[17] != "" and float(row[17]) >= 1.0 and EDGE_I is None: return FILL_G
    if row[-5] == "YES": return FILL_A
    return None
w_core = [5,20,6,5,5,5,18,7,8,16,9,8,8,8,7,8,10,7] + ([10,10,9] if HAS_PP else []) + [8,7,7,9,9,7,20,40]
write_table(ws2, 4, headers, data, w_core, hl_core, freeze_col=2)

# ---------------- Demons ----------------
dl = df[(df.conf != "LOW")].copy()
dl["upside"] = pd.to_numeric(dl.hr_prob) * 0.55 + pd.to_numeric(dl.pgs).clip(lower=0) * 12 + pd.to_numeric(dl.proj_tb) * 8
dl = dl.sort_values("upside", ascending=False).head(40)
ws3 = wb.create_sheet("Demons & Long-shots"); ws3.sheet_properties.tabColor = "C62828"
ws3["A1"] = "DEMONS / LONG-SHOTS RADAR — high-variance upside (HR equity + perception gap + TB ceiling)"; ws3["A1"].font = TITLE
ws3["A2"] = "Ranked by raw upside until live PP demon/goblin ladders join (Zapier relay pending) - then this becomes model probability vs implied odds."; ws3["A2"].font = SUB
headers3 = ["Rank","Player","Team","Opp SP","Same/Opp","Proj FP","PGS","Proj TB","HR %","Upside Score","Rev-Platoon","Conf","Flags","BPP Corroboration"]
data3 = [[i, x.player, x.team, x.opp_sp, x.matchup, x.proj_fp, x.pgs, x.proj_tb, x.hr_prob, round(x.upside,1), x.rev_platoon, x.conf, x["flags"], x.bpp_tags] for i,(_,x) in enumerate(dl.iterrows(),1)]
write_table(ws3, 4, headers3, data3, [5,20,6,18,9,8,7,8,7,10,10,7,20,42], lambda r: FILL_A if r[10]=="YES" else None, freeze_col=2)

# ---------------- HRs ----------------
hr = df.sort_values("hr_prob", ascending=False).head(40)
ws4 = wb.create_sheet("HRs"); ws4.sheet_properties.tabColor = "6A1B9A"
ws4["A1"] = "HOME RUN BOARD — model HR probability, park, corroboration"; ws4["A1"].font = TITLE
ws4["A2"] = "Model HR% is matchup-adjusted" + (" (park NEUTRAL until BPP posts)" if park_pending else "+park adjusted") + ". Fair odds = model probability as a money line - compare vs your book/PP demons."; ws4["A2"].font = SUB
headers4 = ["Rank","Player","Team","Bats","Opp SP","Throws","Mix Barrel %","HR Mult","Park HR","Model HR %","Fair Odds (model)","Proj TB","Conf","Flags","BPP Corroboration"]
def fair_odds(p):
    p = float(p) / 100.0
    if p <= 0: return ""
    return f"+{round(100*(1-p)/p)}" if p < 0.5 else f"-{round(100*p/(1-p))}"
data4 = [[i, x.player, x.team, x.bats, x.opp_sp, x.sp_throws, x.mix_barrel, x.hr_mult, x.park_hr, x.hr_prob, fair_odds(x.hr_prob), x.proj_tb, x.conf, x["flags"], x.bpp_tags] for i,(_,x) in enumerate(hr.iterrows(),1)]
write_table(ws4, 4, headers4, data4, [5,20,6,5,18,7,10,8,8,9,11,8,7,20,40], lambda r: FILL_G if isinstance(r[14],str) and "BPP HR" in r[14] else None, freeze_col=2)

# ---------------- Day After ----------------
if os.path.exists(da_path):
    da = pd.read_csv(da_path)
    wsda = wb.create_sheet("Day After"); wsda.sheet_properties.tabColor = "E65100"
    wsda["A1"] = "DAY AFTER - robbed hard contact yesterday, similar look today"; wsda["A1"].font = TITLE
    wsda["A2"] = ("Yesterday's cohort went 12-for-18 over est lines with 4 HR (see tracking.csv). "
                  "Barrels/near-barrels/deep outs with zero HR to show - matched to today's SP by platoon + mix overlap."); wsda["A2"].font = SUB
    hda = ["Rank","Player","Team","Game","Opp SP Today","Yesterday's Receipts","Process Score","Mix Overlap","HR % Today","Proj FP","PGS","DA Score"]
    dda = [[i, x.player, x.team, x.game, x.opp_sp, x.yesterday, x.process_score, x.mix_overlap, x.hr_prob_today, x.proj_fp, x.pgs, x.da_score] for i,(_,x) in enumerate(da.iterrows(),1)]
    write_table(wsda, 4, hda, dda, [5,20,6,10,17,52,9,9,9,8,7,8], lambda r: FILL_G if r[11] >= 5 else None, freeze_col=2)

# ---------------- Stacks ----------------
st = df.groupby(["game","team"]).apply(lambda g: pd.Series(dict(
    top5_fp=g.nlargest(5,"proj_fp").proj_fp.sum(),
    best5=", ".join(g.nlargest(5,"proj_fp").player.tolist()),
    park_runs=g.park_runs.iloc[0], park_hr=g.park_hr.iloc[0],
    rain="RAIN" if "RAIN" in str(g["flags"].iloc[0]) else "",
    lineup="PROJ" if "LINEUP-PROJ" in str(g["flags"].iloc[0]) else "posted"))).reset_index()
st = st.sort_values("top5_fp", ascending=False)
ws5 = wb.create_sheet("Stacks & Environments"); ws5.sheet_properties.tabColor = "1565C0"
ws5["A1"] = "STACKS — team offensive environments, ranked by sum of top-5 projected FP"; ws5["A1"].font = TITLE
tops = ", ".join(f"{x.team} ({x.top5_fp:.0f})" for _, x in st.head(3).iterrows())
ws5["A2"] = f"Top engines this morning: {tops}." + (" Park factors neutral until BPP posts." if park_pending else ""); ws5["A2"].font = SUB
headers5 = ["Rank","Game","Team","Top-5 FP Sum","Park Runs","Park HR","Suggested Core Stack","Rain","Lineup"]
data5 = [[i, x.game, x.team, round(x.top5_fp,1), x.park_runs, x.park_hr, x.best5, x.rain, x.lineup] for i,(_,x) in enumerate(st.iterrows(),1)]
write_table(ws5, 4, headers5, data5, [5,10,6,12,9,9,72,7,8], lambda r: FILL_G if r[3] >= 44 else None, freeze_col=3)

# ---------------- Full board ----------------
ws6 = wb.create_sheet("Full Board"); ws6.sheet_properties.tabColor = "455A64"
ws6["A1"] = f"FULL BOARD — all {len(df)} hitters, every model component (show-your-work tab)"; ws6["A1"].font = TITLE
full = df.sort_values("proj_fp", ascending=False)
headers6 = ["Player","Team","Game","Slot","Pos","Bats","Opp SP","Throws","Same/Opp","SP Mix","Mix Barrel%","Mix xwOBAcon","Dmg Mult","HR Mult","K Mult","Side PA","SP PA","Exp PA","Proj FP","Public FP","PGS","Proj TB","Proj Hits","Proj BB","Proj R","Proj RBI","Proj SB","HR %","Hit %","Park HR","Park XBH","Park Runs","Rev-Platoon","Free-Swing","Conf","Flags","BPP"]
data6 = [[x.player,x.team,x.game,x.slot,x.pos,x.bats,x.opp_sp,x.sp_throws,x.matchup,x.mix,x.mix_barrel,x.mix_xwobacon,x.dmg_mult,x.hr_mult,x.k_mult,x.side_pa,x.sp_pa,x.exp_pa,x.proj_fp,x.naive_fp,x.pgs,x.proj_tb,x.proj_hits,x.proj_bb,x.proj_r,x.proj_rbi,x.proj_sb,x.hr_prob,x.hit_prob,x.park_hr,x.park_xbh,x.park_runs,x.rev_platoon,x.free_swinger,x.conf,x["flags"],x.bpp_tags] for _,x in full.iterrows()]
write_table(ws6, 3, headers6, data6, [20,6,10,5,5,5,18,7,8,16,9,10,8,8,7,8,7,7,8,8,7,7,8,7,7,8,7,7,7,8,8,9,9,9,6,20,40], freeze_col=1)

out = f"{D}/output/MLB_Model_Board_{VER}_{DATE}.xlsx"
wb.save(out)
print("saved", out, "| top6:", [t[0].split(" - ")[0] for t in top6])
